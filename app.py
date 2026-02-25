import streamlit as st
import anthropic
import fitz  # PyMuPDF
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import os
import io
import base64
import re
from datetime import datetime
from dotenv import load_dotenv
from PIL import Image

# ── Load environment variables ──────────────────────────────────────────────
load_dotenv()

import streamlit as st

ANTHROPIC_API_KEY = st.secrets.get("ANTHROPIC_API_KEY", os.getenv("ANTHROPIC_API_KEY", ""))
GOOGLE_SHEET_ID   = st.secrets.get("GOOGLE_SHEET_ID",   os.getenv("GOOGLE_SHEET_ID", ""))
SHEET_HEADERS = ["Date", "Name", "Topic", "Assessment", "Marks", "Percentage", "Feedback/Remarks"]

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Assessment Grader",
    page_icon="📝",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #1e3a5f 0%, #2d6a9f 100%);
        padding: 2rem;
        border-radius: 12px;
        text-align: center;
        color: white;
        margin-bottom: 2rem;
    }
    .main-header h1 { font-size: 2.2rem; margin: 0; }
    .main-header p  { font-size: 1rem; margin: 0.5rem 0 0; opacity: 0.85; }

    .result-card {
        background: #f8faff;
        border: 1px solid #dce8f8;
        border-left: 4px solid #2d6a9f;
        border-radius: 8px;
        padding: 1.2rem 1.5rem;
        margin: 0.8rem 0;
    }
    .success-card {
        background: #f0fff4;
        border-left: 4px solid #28a745;
        border-radius: 8px;
        padding: 1rem 1.5rem;
        margin: 0.8rem 0;
    }
    .warning-card {
        background: #fffbf0;
        border-left: 4px solid #ffc107;
        border-radius: 8px;
        padding: 1rem 1.5rem;
        margin: 0.8rem 0;
    }
    .error-card {
        background: #fff5f5;
        border-left: 4px solid #dc3545;
        border-radius: 8px;
        padding: 1rem 1.5rem;
        margin: 0.8rem 0;
    }
    .metric-row { display: flex; gap: 1rem; margin: 1rem 0; }
    .metric-box {
        flex: 1;
        background: white;
        border: 1px solid #dce8f8;
        border-radius: 8px;
        padding: 1rem;
        text-align: center;
    }
    .metric-box .value { font-size: 2rem; font-weight: bold; color: #2d6a9f; }
    .metric-box .label { font-size: 0.85rem; color: #666; }
    .stButton > button {
        width: 100%;
        background: linear-gradient(135deg, #1e3a5f, #2d6a9f);
        color: white;
        border: none;
        padding: 0.6rem 1.5rem;
        border-radius: 8px;
        font-size: 1rem;
        font-weight: 600;
    }
    .stButton > button:hover { opacity: 0.9; }
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════════
#  HELPERS
# ════════════════════════════════════════════════════════════════════

def pdf_to_base64_images(pdf_bytes: bytes) -> list[str]:
    """Convert every page of a PDF to a base64-encoded PNG string."""
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    images = []
    for page in doc:
        mat  = fitz.Matrix(2.0, 2.0)          # 2x zoom → higher quality
        pix  = page.get_pixmap(matrix=mat, alpha=False)
        png  = pix.tobytes("png")
        images.append(base64.standard_b64encode(png).decode())
    doc.close()
    return images


def extract_and_grade_with_claude(images_b64: list[str], api_key: str) -> dict:
    """
    Send all PDF pages to Claude Vision.
    Claude extracts meta-data AND grades the answers.
    Returns a structured dict.
    """
    client = anthropic.Anthropic(api_key=api_key)

    content = []
    for img in images_b64:
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": "image/png", "data": img}
        })

    content.append({
        "type": "text",
        "text": """You are an expert technical assessor for an IT training organization.

You will receive images of a scanned intern assessment answer sheet (could be handwritten or printed).

Your tasks:
1. Extract the following metadata from the document:
   - date (written on the paper — use exactly as written, format as DD-MM-YYYY if possible)
   - intern_name (full name of the intern)
   - topic (e.g., "MSSQL", "Python", "SQL", etc.)
   - assessment_number (e.g., "Assessment 1", "Assessment 2", etc.)

2. Extract every question and the intern's answer.

3. For each question:
   - Determine the correct answer using your knowledge
   - Compare with the intern's answer
   - Award marks: 1 mark per question (partial credit: 0.5 if partially correct)
   - Write brief feedback

4. Calculate:
   - total_marks: sum of marks awarded
   - max_marks: total number of questions
   - percentage: (total_marks / max_marks) * 100

5. Write overall_feedback summarizing performance (2-3 sentences).

Respond ONLY with a valid JSON object in this exact format:
{
  "date": "DD-MM-YYYY",
  "intern_name": "Full Name",
  "topic": "Topic Name",
  "assessment_number": "Assessment N",
  "questions": [
    {
      "q_no": 1,
      "question": "Question text",
      "intern_answer": "What intern wrote",
      "correct_answer": "Correct answer",
      "marks_awarded": 1,
      "max_marks": 1,
      "feedback": "Brief comment"
    }
  ],
  "total_marks": 0,
  "max_marks": 0,
  "percentage": 0.0,
  "overall_feedback": "Overall performance summary."
}

If you cannot read a field, use "Unknown" for strings and 0 for numbers.
Return ONLY the JSON, no markdown, no explanation."""
    })

    response = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=4096,
        messages=[{"role": "user", "content": content}]
    )

    raw = response.content[0].text.strip()
    # Strip markdown code fences if present
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)


# ── Google Sheets helpers ────────────────────────────────────────────────────

def get_gsheet(creds_path: str, sheet_id: str):
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds  = Credentials.from_service_account_file(creds_path, scopes=scopes)
    gc     = gspread.authorize(creds)
    return gc.open_by_key(sheet_id).sheet1


def ensure_headers(ws):
    existing = ws.row_values(1)
    if existing != SHEET_HEADERS:
        ws.clear()
        ws.append_row(SHEET_HEADERS)
        # Bold header
        ws.format("A1:G1", {
            "textFormat": {"bold": True},
            "backgroundColor": {"red": 0.18, "green": 0.23, "blue": 0.37}
        })


def is_duplicate(ws, name: str, assessment: str) -> bool:
    records = ws.get_all_records()
    for row in records:
        if (str(row.get("Name", "")).strip().lower() == name.strip().lower() and
                str(row.get("Assessment", "")).strip().lower() == assessment.strip().lower()):
            return True
    return False


def append_to_sheet(ws, row_data: dict):
    ws.append_row([
        row_data["date"],
        row_data["intern_name"],
        row_data["topic"],
        row_data["assessment_number"],
        f"{row_data['total_marks']}/{row_data['max_marks']}",
        f"{row_data['percentage']:.1f}%",
        row_data["overall_feedback"],
    ])


# ── Excel download helper ────────────────────────────────────────────────────

def build_excel(result: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assessment Result"

    # Styles
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1E3A5F")
    center       = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin         = Side(border_style="thin", color="AAAAAA")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Summary section ──────────────────────────────────────────────
    summary_headers = SHEET_HEADERS
    for col, h in enumerate(summary_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    data_row = [
        result["date"],
        result["intern_name"],
        result["topic"],
        result["assessment_number"],
        f"{result['total_marks']}/{result['max_marks']}",
        f"{result['percentage']:.1f}%",
        result["overall_feedback"],
    ]
    for col, val in enumerate(data_row, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.alignment = center
        cell.border    = border

    # ── Per-question section ─────────────────────────────────────────
    ws.cell(row=4, column=1, value="Question-wise Breakdown").font = Font(bold=True, size=12)
    q_headers = ["Q.No", "Question", "Intern's Answer", "Correct Answer", "Marks", "Feedback"]
    q_fill    = PatternFill("solid", fgColor="2D6A9F")
    for col, h in enumerate(q_headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = q_fill
        cell.alignment = center
        cell.border    = border

    for i, q in enumerate(result["questions"], 6):
        row_vals = [
            q.get("q_no", i - 5),
            q.get("question", ""),
            q.get("intern_answer", ""),
            q.get("correct_answer", ""),
            f"{q.get('marks_awarded', 0)}/{q.get('max_marks', 1)}",
            q.get("feedback", ""),
        ]
        bg = "F0FFF4" if q.get("marks_awarded", 0) >= q.get("max_marks", 1) else \
             "FFF5F5" if q.get("marks_awarded", 0) == 0 else "FFFBF0"
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = border
            cell.fill      = PatternFill("solid", fgColor=bg)

    # Column widths
    col_widths = [8, 40, 35, 35, 10, 40]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 30

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ════════════════════════════════════════════════════════════════════
#  SIDEBAR — CONFIGURATION
# ════════════════════════════════════════════════════════════════════

with st.sidebar:
    st.image("https://img.icons8.com/color/96/graduation-cap.png", width=80)
    st.title("⚙️ Configuration")

    st.subheader("🔑 API Keys")
    api_key_input = st.text_input(
        "Anthropic API Key",
        value=ANTHROPIC_API_KEY,
        type="password",
        help="Get your key at console.anthropic.com",
    )

    st.subheader("📊 Google Sheets")
    sheet_id_input = st.text_input(
        "Google Sheet ID",
        value=GOOGLE_SHEET_ID,
        help="Found in the Sheet URL: /spreadsheets/d/<SHEET_ID>/",
    )
    creds_file = st.file_uploader(
        "Upload credentials.json",
        type="json",
        help="Google Service Account JSON key file",
    )

    # Save uploaded credentials temporarily
    creds_path = GOOGLE_CREDS_PATH
    if creds_file:
        creds_path = "/tmp/uploaded_credentials.json"
        with open(creds_path, "wb") as f:
            f.write(creds_file.read())
        st.success("✅ credentials.json loaded")

    st.divider()
    st.markdown("### 📌 How to use")
    st.markdown("""
1. Fill in your **API key** and **Sheet ID**
2. Upload your Google **credentials.json**
3. Upload a **scanned PDF** assessment
4. Click **Grade Assessment**
5. Download the **Excel report**
""")
    st.divider()
    st.caption("Built with ❤️ using Claude AI + Streamlit")


# ════════════════════════════════════════════════════════════════════
#  MAIN — HEADER
# ════════════════════════════════════════════════════════════════════

st.markdown("""
<div class="main-header">
    <h1>📝 Assessment Auto-Grader</h1>
    <p>Upload a scanned PDF → AI extracts & grades → Saves to Google Sheets</p>
</div>
""", unsafe_allow_html=True)

# ── Upload section ───────────────────────────────────────────────────────────
col1, col2 = st.columns([2, 1])

with col1:
    uploaded_pdf = st.file_uploader(
        "📂 Upload Scanned Assessment PDF",
        type=["pdf"],
        help="Scanned or digital PDF of the intern's assessment answer sheet",
    )

with col2:
    st.markdown("### ℹ️ Requirements")
    st.info("""
- PDF can be scanned / handwritten
- Should contain: **Name**, **Date**, **Topic**, **Assessment No.**, **Q&A**
- Duplicate entries are automatically blocked
""")

# ── Grade button ─────────────────────────────────────────────────────────────
if uploaded_pdf:
    st.markdown("---")
    if st.button("🚀 Grade This Assessment", use_container_width=True):

        # Validate config
        missing = []
        if not api_key_input:  missing.append("Anthropic API Key")
        if not sheet_id_input: missing.append("Google Sheet ID")
        if not os.path.exists(creds_path): missing.append("Google credentials.json")

        if missing:
            st.markdown(f'<div class="error-card">❌ Missing configuration: <b>{", ".join(missing)}</b></div>',
                        unsafe_allow_html=True)
            st.stop()

        pdf_bytes = uploaded_pdf.read()

        # Step 1 — Convert PDF to images
        with st.spinner("📄 Converting PDF pages to images…"):
            try:
                images_b64 = pdf_to_base64_images(pdf_bytes)
                st.success(f"✅ Converted {len(images_b64)} page(s)")
            except Exception as e:
                st.error(f"Failed to read PDF: {e}")
                st.stop()

        # Step 2 — Claude AI extraction + grading
        with st.spinner("🤖 Claude AI is reading and grading the assessment… (this may take 20-40 sec)"):
            try:
                result = extract_and_grade_with_claude(images_b64, api_key_input)
            except json.JSONDecodeError as e:
                st.error(f"AI returned invalid JSON. Try re-uploading a clearer scan. Details: {e}")
                st.stop()
            except Exception as e:
                st.error(f"Claude API error: {e}")
                st.stop()

        # Step 3 — Display extracted info
        st.markdown("## 📋 Extracted Information")
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("👤 Name",       result.get("intern_name", "Unknown"))
        m2.metric("📅 Date",       result.get("date", "Unknown"))
        m3.metric("📚 Topic",      result.get("topic", "Unknown"))
        m4.metric("🔢 Assessment", result.get("assessment_number", "Unknown"))

        # Step 4 — Scoring
        st.markdown("## 📊 Score")
        s1, s2, s3 = st.columns(3)
        s1.metric("Marks Awarded",  f"{result.get('total_marks', 0)} / {result.get('max_marks', 0)}")
        s2.metric("Percentage",     f"{result.get('percentage', 0):.1f}%")
        pct = result.get("percentage", 0)
        grade = "A+" if pct >= 90 else "A" if pct >= 80 else "B" if pct >= 70 else \
                "C" if pct >= 60 else "D" if pct >= 50 else "F"
        s3.metric("Grade", grade)

        # Overall feedback
        st.markdown("### 💬 Overall Feedback")
        st.info(result.get("overall_feedback", "No feedback generated."))

        # Step 5 — Question breakdown
        st.markdown("## 📝 Question-wise Breakdown")
        questions = result.get("questions", [])
        if questions:
            for q in questions:
                awarded = q.get("marks_awarded", 0)
                maximum = q.get("max_marks", 1)
                icon    = "✅" if awarded >= maximum else "⚠️" if awarded > 0 else "❌"
                with st.expander(f"{icon} Q{q.get('q_no', '?')} — {q.get('question', '')[:80]}... | {awarded}/{maximum} marks"):
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown("**Intern's Answer:**")
                        st.write(q.get("intern_answer", "—"))
                    with c2:
                        st.markdown("**Correct Answer:**")
                        st.write(q.get("correct_answer", "—"))
                    st.markdown(f"**Feedback:** {q.get('feedback', '—')}")
        else:
            st.warning("No questions were extracted.")

        # Step 6 — Google Sheets
        st.markdown("## 💾 Saving to Google Sheets")
        try:
            ws = get_gsheet(creds_path, sheet_id_input)
            ensure_headers(ws)

            name       = result.get("intern_name", "Unknown")
            assessment = result.get("assessment_number", "Unknown")

            if is_duplicate(ws, name, assessment):
                st.markdown(
                    f'<div class="warning-card">⚠️ <b>Duplicate detected!</b> '
                    f'<i>{name}</i> already has a record for <i>{assessment}</i>. '
                    f'This record was NOT saved to Google Sheets.</div>',
                    unsafe_allow_html=True,
                )
            else:
                append_to_sheet(ws, result)
                st.markdown(
                    '<div class="success-card">✅ Result saved to Google Sheets successfully!</div>',
                    unsafe_allow_html=True,
                )
        except FileNotFoundError:
            st.error("credentials.json not found. Please upload it in the sidebar.")
        except Exception as e:
            st.error(f"Google Sheets error: {e}")

        # Step 7 — Excel download
        st.markdown("## ⬇️ Download Excel Report")
        excel_bytes = build_excel(result)
        safe_name   = re.sub(r"[^a-zA-Z0-9_]", "_", result.get("intern_name", "assessment"))
        filename    = f"{safe_name}_{result.get('assessment_number', 'result').replace(' ', '_')}.xlsx"
        st.download_button(
            label="📥 Download Detailed Excel Report",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

# ── Google Sheet viewer ──────────────────────────────────────────────────────
st.markdown("---")
st.markdown("## 📊 View All Records")
if st.button("🔄 Load Records from Google Sheets", use_container_width=True):
    try:
        ws      = get_gsheet(creds_path, sheet_id_input)
        records = ws.get_all_records()
        if records:
            df = pd.DataFrame(records)
            st.dataframe(df, use_container_width=True, height=400)
            st.caption(f"Total records: {len(df)}")

            # Download full sheet as Excel
            out = io.BytesIO()
            df.to_excel(out, index=False)
            st.download_button(
                "📥 Download All Records as Excel",
                data=out.getvalue(),
                file_name="all_assessments.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.info("No records found in the sheet yet.")
    except Exception as e:
        st.error(f"Could not load records: {e}")

# ── Footer ───────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    "<p style='text-align:center; color:#888; font-size:0.85rem;'>"
    "Assessment Auto-Grader · Powered by Claude AI · Built with Streamlit"
    "</p>",
    unsafe_allow_html=True,
)
